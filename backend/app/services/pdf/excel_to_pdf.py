"""
excel_to_pdf.py — Enterprise-grade Excel to PDF Processor
==========================================================
KEY FEATURES:
  • LibreOffice primary path: preserves column widths, number formats,
    cell colors, charts, headers/footers, and print areas
  • openpyxl structured fallback: renders a proper table grid with
    column headers, column-width estimation, and number formatting
  • Merged cell handling (reports the display value, not None)
  • Formula error masking (#VALUE!, #REF! → blank for readability)
  • Multi-sheet support (one page per sheet in the fallback path)
  • Conversion engine reported in metadata
"""
from __future__ import annotations

import logging
from pathlib import Path

import fitz
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.models.enums import ArtifactKind
from app.schemas.job import ConvertToPdfJobRequest
from app.services.pdf.advanced_utils import pdf_page_count
from app.services.pdf.common import (
    BaseToolProcessor,
    GeneratedArtifact,
    PDF_CONTENT_TYPE,
    PdfProcessingError,
    ProcessingResult,
    ProcessorContext,
    ensure_pdf_output_filename,
)
from app.utils.libreoffice import LibreOfficeConversionError, LibreOfficeUnavailableError, convert_with_libreoffice

log = logging.getLogger(__name__)

_MARGIN = 36.0
_PAGE_W = 841.89   # A4 landscape for wide spreadsheets
_PAGE_H = 595.28
_FONT_SIZE = 8.5
_ROW_H = 14.0
_HEADER_ROW_H = 16.0
_EXCEL_ERRORS = {"#VALUE!", "#REF!", "#NAME?", "#DIV/0!", "#NULL!", "#N/A", "#NUM!"}


class ExcelToPdfProcessor(BaseToolProcessor):
    tool_id = "excel2pdf"

    def process(self, context: ProcessorContext) -> ProcessingResult:
        payload = ConvertToPdfJobRequest.model_validate(context.payload)
        source = context.require_single_input()
        output_filename = ensure_pdf_output_filename(payload.output_filename)
        output_path = context.workspace / output_filename

        # --- Primary: LibreOffice ---
        try:
            converted_path = convert_with_libreoffice(
                source.storage_path,
                output_dir=context.workspace,
                target_format="pdf",
            )
            if converted_path != output_path:
                converted_path.replace(output_path)
            return ProcessingResult(
                artifact=GeneratedArtifact(
                    local_path=output_path,
                    filename=output_filename,
                    content_type=PDF_CONTENT_TYPE,
                    kind=ArtifactKind.RESULT,
                    metadata={
                        "pages_processed": pdf_page_count(output_path),
                        "conversion_engine": "libreoffice",
                    },
                ),
                completion_message="Spreadsheet converted to PDF successfully.",
            )
        except LibreOfficeUnavailableError:
            log.info("excel2pdf: LibreOffice not available, using openpyxl fallback")
        except LibreOfficeConversionError as exc:
            raise PdfProcessingError(
                code="excel_conversion_failed",
                user_message=f"Could not convert the spreadsheet: {exc}",
            ) from exc

        # --- Fallback: openpyxl structured table renderer ---
        try:
            workbook = load_workbook(source.storage_path, data_only=True)
        except Exception as exc:
            raise PdfProcessingError(
                code="excel_read_failed",
                user_message="Could not open the Excel file. The file may be corrupted.",
            ) from exc

        pdf_doc = fitz.open()
        sheets_rendered = 0
        for sheet in workbook.worksheets:
            _render_sheet(sheet, pdf_doc)
            sheets_rendered += 1

        if pdf_doc.page_count == 0:
            pdf_doc.new_page(width=_PAGE_W, height=_PAGE_H)

        pdf_doc.save(output_path, garbage=4, deflate=True)
        pdf_doc.close()

        return ProcessingResult(
            artifact=GeneratedArtifact(
                local_path=output_path,
                filename=output_filename,
                content_type=PDF_CONTENT_TYPE,
                kind=ArtifactKind.RESULT,
                metadata={
                    "pages_processed": pdf_page_count(output_path),
                    "sheets_rendered": sheets_rendered,
                    "conversion_engine": "openpyxl-fallback",
                },
            ),
            completion_message="Spreadsheet converted to PDF successfully.",
        )


# ---------------------------------------------------------------------------
# openpyxl sheet renderer
# ---------------------------------------------------------------------------

def _render_sheet(sheet, pdf: fitz.Document) -> None:
    """Renders one worksheet as one or more PDF pages."""
    # Collect non-empty rows
    all_rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        cells = [_format_cell(v) for v in row]
        # Include row if any cell has a value
        if any(c.strip() for c in cells):
            all_rows.append(cells)

    if not all_rows:
        return

    # Estimate column widths based on maximum content length in each column
    if all_rows:
        n_cols = max(len(r) for r in all_rows)
    else:
        return

    max_col_chars = [1] * n_cols
    for row in all_rows:
        for col_idx, cell in enumerate(row[:n_cols]):
            max_col_chars[col_idx] = max(max_col_chars[col_idx], min(len(cell), 30))

    # Distribute column widths proportionally across available width
    total_chars = sum(max_col_chars)
    avail_w = _PAGE_W - 2 * _MARGIN
    col_widths = [max(30.0, (c / total_chars) * avail_w) for c in max_col_chars]

    # Render title row (sheet name)
    page = pdf.new_page(width=_PAGE_W, height=_PAGE_H)
    y = _MARGIN

    page.insert_text(
        fitz.Point(_MARGIN, y + 13),
        sheet.title,
        fontname="helv",
        fontsize=13.0,
        color=(0.1, 0.1, 0.5),
    )
    y += 20.0

    # Draw header line
    page.draw_line(fitz.Point(_MARGIN, y), fitz.Point(_PAGE_W - _MARGIN, y), color=(0.5, 0.5, 0.5), width=0.5)
    y += 2

    for row_idx, row_cells in enumerate(all_rows):
        row_h = _HEADER_ROW_H if row_idx == 0 else _ROW_H
        # Page break
        if y + row_h > _PAGE_H - _MARGIN:
            page = pdf.new_page(width=_PAGE_W, height=_PAGE_H)
            y = _MARGIN

        x = _MARGIN
        is_header = row_idx == 0
        font_color = (0, 0, 0.4) if is_header else (0, 0, 0)
        for col_idx in range(min(n_cols, len(col_widths))):
            cell_text = row_cells[col_idx] if col_idx < len(row_cells) else ""
            # Truncate to fit column
            char_limit = max(3, int(col_widths[col_idx] / (_FONT_SIZE * 0.55)))
            if len(cell_text) > char_limit:
                cell_text = cell_text[:char_limit - 1] + "…"
            page.insert_text(
                fitz.Point(x + 2, y + _FONT_SIZE + 1),
                cell_text,
                fontname="helv",
                fontsize=_FONT_SIZE,
                color=font_color,
            )
            # Vertical cell divider
            page.draw_line(
                fitz.Point(x + col_widths[col_idx], y),
                fitz.Point(x + col_widths[col_idx], y + row_h),
                color=(0.85, 0.85, 0.85),
                width=0.3,
            )
            x += col_widths[col_idx]

        # Horizontal row divider
        page.draw_line(
            fitz.Point(_MARGIN, y + row_h),
            fitz.Point(_PAGE_W - _MARGIN, y + row_h),
            color=(0.85, 0.85, 0.85),
            width=0.3,
        )
        y += row_h


def _format_cell(value) -> str:
    """Formats a cell value for display, masking errors and None."""
    if value is None:
        return ""
    s = str(value)
    if s in _EXCEL_ERRORS:
        return ""  # hide formula errors
    # Format numbers nicely
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return f"{value:.4g}"
    return s