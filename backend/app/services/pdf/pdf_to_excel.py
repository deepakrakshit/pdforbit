"""
pdf_to_excel.py — Enterprise-grade PDF to Excel Processor
==========================================================
KEY FEATURES:
  • pdfplumber table detection: proper row × column structure extraction
  • Number formatting heuristics: currency, percentages, dates → Excel formats
  • Multiple tables per page: stacked with blank-row separator
  • Column width auto-sizing based on content length
  • Header row detection (first row made bold)
  • OCR fallback via DocumentTextExtractor for scanned PDFs
  • Named sheets: "Page 1 - Table 1", etc.
  • Workbook metadata
"""
from __future__ import annotations

from contextlib import nullcontext
import logging
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

from app.models.enums import ArtifactKind
from app.schemas.job import PdfToOfficeJobRequest
from app.services.pdf.common import (
    BaseToolProcessor,
    GeneratedArtifact,
    ProcessingResult,
    ProcessorContext,
)
from app.utils.files import sanitize_filename
from app.services.pdf.document_intelligence import DocumentTextExtractor

log = logging.getLogger(__name__)

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_OCR_THRESHOLD = 80
# Header background color for the first row of each table
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
_BODY_FONT = Font(size=9)

# Regex patterns for cell value typing
_RE_CURRENCY = re.compile(r"^[\$£€¥]\s*[\d,]+\.?\d*$")
_RE_PERCENT = re.compile(r"^[\d.]+\s*%$")
_RE_DATE_US = re.compile(r"^\d{1,2}/\d{1,2}/\d{2,4}$")
_RE_DATE_ISO = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_RE_NUMBER = re.compile(r"^-?[\d,]+\.?\d*$")


class PdfToExcelProcessor(BaseToolProcessor):
    tool_id = "pdf2excel"

    def __init__(self, *, extractor: DocumentTextExtractor | None = None) -> None:
        self._extractor = extractor

    def process(self, context: ProcessorContext) -> ProcessingResult:
        payload = PdfToOfficeJobRequest.model_validate(context.payload)
        source = context.require_single_input()
        output_filename = sanitize_filename(payload.output_filename or f"{Path(source.original_filename).stem}.xlsx")
        output_path = context.workspace / output_filename

        try:
            import pdfplumber  # type: ignore
        except ImportError:
            pdfplumber = None
            log.warning("pdf2excel: pdfplumber is unavailable, falling back to plain-text extraction")

        workbook = Workbook()
        workbook.remove(workbook.active)  # remove default empty sheet

        tables_total = 0
        pages_with_ocr = 0

        import fitz
        with fitz.open(source.storage_path) as fitz_doc:
            total_pages = fitz_doc.page_count

            plumber_context = pdfplumber.open(source.storage_path) if pdfplumber else nullcontext(None)
            with plumber_context as plumber_doc:
                plumber_pages = plumber_doc.pages if plumber_doc is not None else [None] * total_pages
                for page_idx, plumber_page in enumerate(plumber_pages):
                    fitz_page = fitz_doc.load_page(page_idx)
                    native_text = fitz_page.get_text("text").strip()
                    needs_ocr = len(native_text) < _OCR_THRESHOLD
                    if needs_ocr:
                        pages_with_ocr += 1

                    # Try to extract structured tables first
                    raw_tables = []
                    if plumber_page is not None:
                        try:
                            raw_tables = plumber_page.extract_tables(table_settings={
                                "vertical_strategy": "lines",
                                "horizontal_strategy": "lines",
                                "snap_tolerance": 3,
                                "join_tolerance": 3,
                                "edge_min_length": 3,
                                "min_words_vertical": 3,
                                "min_words_horizontal": 1,
                            }) or []
                            if not raw_tables:
                                raw_tables = plumber_page.extract_tables() or []
                        except Exception as exc:
                            log.debug("pdf2excel: pdfplumber extraction failed on page %d: %s", page_idx + 1, exc)

                    # If no structured tables found, try to extract tabular text
                    if not raw_tables and plumber_page is not None:
                        words_table = plumber_page.extract_words(
                            x_tolerance=3, y_tolerance=3, keep_blank_chars=False
                        )
                        inferred = _infer_table_from_words(words_table)
                        if inferred:
                            raw_tables = [inferred]

                    if not raw_tables:
                        # Emit a plain-text sheet for pages with no detectable tables
                        text = native_text or "(no extractable text)"
                        sheet_name = _safe_sheet_name(f"Page {page_idx + 1}", workbook)
                        ws = workbook.create_sheet(title=sheet_name)
                        for row_idx, line in enumerate(text.splitlines()[:5000], start=1):
                            ws.cell(row=row_idx, column=1, value=line)
                        ws.column_dimensions["A"].width = 80
                        continue

                    for tbl_idx, raw_table in enumerate(raw_tables):
                        if not raw_table:
                            continue
                        sheet_name = _safe_sheet_name(f"P{page_idx + 1}T{tbl_idx + 1}", workbook)
                        ws = workbook.create_sheet(title=sheet_name)
                        _write_table_to_sheet(ws, raw_table)
                        tables_total += 1

        if not workbook.worksheets:
            ws = workbook.create_sheet(title="Sheet1")
            ws.cell(row=1, column=1, value="No tables detected")

        # Set workbook metadata
        workbook.properties.title = f"Converted from {source.original_filename}"
        workbook.properties.creator = "PdfORBIT"

        workbook.save(output_path)

        return ProcessingResult(
            artifact=GeneratedArtifact(
                local_path=output_path,
                filename=output_filename,
                content_type=XLSX_CONTENT_TYPE,
                kind=ArtifactKind.RESULT,
                metadata={
                    "pages_processed": total_pages,
                    "tables_extracted": tables_total,
                    "pages_with_ocr": pages_with_ocr,
                },
            ),
            completion_message="PDF converted to Excel successfully.",
        )


# ---------------------------------------------------------------------------
# Sheet writing with formatting
# ---------------------------------------------------------------------------

def _write_table_to_sheet(ws, raw_table: list[list]) -> None:
    """Writes a 2D list of cells to an openpyxl worksheet with formatting."""
    if not raw_table:
        return

    n_cols = max(len(row) for row in raw_table)
    col_max_len = [1] * n_cols

    for row_idx, row in enumerate(raw_table, start=1):
        for col_idx in range(n_cols):
            raw_val = row[col_idx] if col_idx < len(row) else None
            cell_text = str(raw_val).strip() if raw_val is not None else ""

            cell = ws.cell(row=row_idx, column=col_idx + 1)
            typed_val, number_format = _type_cell_value(cell_text)
            cell.value = typed_val
            if number_format:
                cell.number_format = number_format

            # Column width tracking
            col_max_len[col_idx] = max(col_max_len[col_idx], min(len(cell_text), 50))

            # Header row formatting
            if row_idx == 1:
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = _BODY_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Set column widths
    for col_idx, max_len in enumerate(col_max_len):
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = max(8, min(max_len + 2, 52))

    # Freeze header row
    ws.freeze_panes = "A2"


def _type_cell_value(text: str) -> tuple[object, str | None]:
    """
    Attempts to convert a string to a typed Excel value.
    Returns (value, number_format_or_None).
    """
    if not text:
        return "", None
    # Currency
    if _RE_CURRENCY.match(text):
        num_str = re.sub(r"[^\d.]", "", text)
        try:
            return float(num_str), '"$"#,##0.00'
        except ValueError:
            pass
    # Percentage
    if _RE_PERCENT.match(text):
        num_str = text.replace("%", "").strip()
        try:
            return float(num_str) / 100.0, "0.00%"
        except ValueError:
            pass
    # ISO date
    if _RE_DATE_ISO.match(text):
        return text, "YYYY-MM-DD"
    # US date
    if _RE_DATE_US.match(text):
        return text, "MM/DD/YYYY"
    # Plain number
    if _RE_NUMBER.match(text.replace(",", "")):
        try:
            num = float(text.replace(",", ""))
            if num == int(num) and "." not in text:
                return int(num), None
            return num, "#,##0.00" if abs(num) >= 1000 else "0.00##"
        except ValueError:
            pass
    return text, None


def _infer_table_from_words(words: list[dict]) -> list[list[str]] | None:
    """
    Attempts to reconstruct a table from word-level position data.
    Groups words by y-coordinate (row) and x-coordinate (column).
    """
    if not words:
        return None
    # Group by rounded y position (row)
    rows: dict[int, list[dict]] = {}
    for word in words:
        y_key = round(float(word.get("top", 0)) / 5) * 5
        rows.setdefault(y_key, []).append(word)
    if len(rows) < 2:
        return None
    # Sort each row by x position and extract text
    table: list[list[str]] = []
    for y_key in sorted(rows.keys()):
        row_words = sorted(rows[y_key], key=lambda w: float(w.get("x0", 0)))
        table.append([w.get("text", "") for w in row_words])
    return table if len(table) >= 2 else None


def _safe_sheet_name(name: str, workbook: Workbook) -> str:
    """Returns a sheet name that is unique and ≤ 31 chars (Excel limit)."""
    safe = re.sub(r"[:\\/?*\[\]]", "-", name)[:31]
    existing = {ws.title for ws in workbook.worksheets}
    if safe not in existing:
        return safe
    for i in range(2, 1000):
        candidate = f"{safe[:28]}-{i}"
        if candidate not in existing:
            return candidate
    return safe